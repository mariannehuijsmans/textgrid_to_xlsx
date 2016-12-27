import os
from openpyxl import Workbook


"""
This code takes a tab-separated txt file, output of a textgrid,
where the dependency between tiers is lost, and recreates the dependencies
using the time stamps and stores the information in a dictionary
where entries are lists of dictionary entries. It then creates an excel spreadsheet
and writes the entries that belong to one audio segment all in one row.

The code is designed for textgrid that has tiers for the name of the audio file
(minus the prefix) 'file', the transcription of the word 'words', a morphological breakdown 'morphemes',
a gloss 'gloss', a translation 'translation', a second translation 'translation2',
notes 'notes', and any edits that need to be done for the audio files 'edits'.
The code can easily be altered for different tiers and tier names. This will
effect both the creation of the lemmalist storing the information and the
writing of the excel spreadsheet, however.

To use the program, 1) change the name of the txt file, 2) change the prefix
to the prefix the audio files will have, 3) change the speaker, recorder,
elicitor, place values, 4) change the name of the output file
"""

path = os.path.join(os.getcwd(), '160810-002.txt')

prefix = 'aug102016_EP_'

speaker = 'Elsie Paul'

recorder = 'Marianne Huijsmans'

elicitor = 'Marianne Huijsmans'

place = 'Elsie\'s house'

output = "aug102016_EP.xlsx"

lines = []
with open(path, mode = 'r', encoding = 'utf-8') as file:
    for line in file:
        line = line.strip()
        line = line.split('\t')
        lines.append(line)

lemmalist = {}


for line in lines:
    if line[0] == 'words':
        word = line[4]
        lemmalist[word] = []
        begintime = line[2]
        lemmalist[word].append(begintime)
        
for word in lemmalist:
    for line in lines:
        if len(line) >= 5:
            if lemmalist[word][0] == line[2]:
                if line[0] == 'file':
                    audio = {}
                    line[4] = prefix+line[4]
                    audio['Audio'] = line[4]
                    lemmalist[word].append(audio)
                if line[0] == 'morphemes':
                    morphemes = {}
                    morphemes['Morphemes']= line[4]
                    lemmalist[word].append(morphemes)
                if line[0] == 'gloss':
                    gloss = {}
                    gloss['Gloss']= line[4]
                    lemmalist[word].append(gloss)
                if line[0] == 'translation':
                    translation = {}
                    translation['Translation']= line[4]
                    lemmalist[word].append(translation)
                if line[0] == 'translation2':
                    translation2 = {}
                    translation2['Translation2']= line[4]
                    lemmalist[word].append(translation2)
                if line[0] == 'notes':
                    notes = {}
                    notes['Notes'] = line[4]
                    lemmalist[word].append(notes)
                if line[0] == 'edits':
                    edits = {}
                    edits['Edits']= line[4]
                    lemmalist[word].append(edits)
        else:           
            continue


wb = Workbook()
ws = wb.active           

for j,heading in enumerate(['Audio', 'Speaker', 'Elicitor', 'Recorder', 'Place','Words','Morphemes', 'Gloss', 'Translation', 'Translation2', 'Notes', 'Edits']):
    ws.cell(row=1, column=j+1).value=heading
    
for i, entry in enumerate(lemmalist):
    ws.cell(row=i+2, column=2).value=speaker
    ws.cell(row=i+2, column=3).value=elicitor
    ws.cell(row=i+2, column=4).value=recorder
    ws.cell(row=i+2, column=5).value=place
    ws.cell(row=i+2, column=6).value=entry
    for item in lemmalist[entry]:
        for key in item:
            if key == 'Audio':
                ws.cell(row=i+2, column=1).value= item[key]
            if key == 'Morphemes':
                ws.cell(row=i+2, column=7).value= item[key]
            if key == 'Gloss':
                ws.cell(row=i+2, column=8).value= item[key]
            if key == 'Translation':
                ws.cell(row=i+2, column=9).value= item[key]
            if key == 'Translation2':
                ws.cell(row=i+2, column=10).value= item[key]
            if key == 'Notes':
                ws.cell(row=i+2, column=11).value= item[key]
            if key == 'Edits':
                ws.cell(row=i+2, column=12).value= item[key]

            
wb.save(output)
                

        
        
            
                
